"""
CSV / XLSX extractor.

Capabilities:
- Reads CSV files (auto-detecting delimiter)
- Reads all sheets from XLSX workbooks
- Detects header row even when not at row 1
- Preserves source row indices for traceability
- Returns ExtractionResult with headers and dicts
"""

from __future__ import annotations

import csv
import io
import logging
import os
from typing import Any, Dict, List, Optional, Tuple

from ..contracts import CanonicalType, ExtractionResult
from ..row_locator import extract_data_rows, locate_header_row

logger = logging.getLogger(__name__)


def _detect_csv_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def read_csv(
    path: str,
    canonical_type: CanonicalType = CanonicalType.UNKNOWN,
    known_aliases: Optional[Dict[str, str]] = None,
) -> ExtractionResult:
    notes: List[str] = []
    try:
        with open(path, encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096)
        delimiter = _detect_csv_delimiter(sample)

        with open(path, encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            raw_rows: List[List[Any]] = list(reader)

        header_idx, headers = locate_header_row(raw_rows, known_aliases=known_aliases)
        if header_idx > 0:
            notes.append(f"Header detected at row {header_idx + 1} (1-based).")

        data_rows = extract_data_rows(raw_rows, header_idx, headers)
        return ExtractionResult(
            source_path=path,
            source_type=canonical_type,
            headers=headers,
            rows=data_rows,
            header_row_index=header_idx,
            sheet_name=None,
            extraction_notes=notes,
        )
    except Exception as e:
        logger.error(f"CSV read failed for {path}: {e}")
        return ExtractionResult(
            source_path=path,
            source_type=canonical_type,
            headers=[],
            rows=[],
            header_row_index=None,
            sheet_name=None,
            extraction_notes=[f"ERROR: {e}"],
        )


def read_excel(
    path: str,
    canonical_type: CanonicalType = CanonicalType.UNKNOWN,
    known_aliases: Optional[Dict[str, str]] = None,
    target_sheet: Optional[str] = None,
) -> List[ExtractionResult]:
    """
    Read all sheets (or a specific sheet) from an XLSX file.
    Returns one ExtractionResult per sheet.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        logger.error("openpyxl not installed. Cannot read XLSX. pip install openpyxl")
        return [ExtractionResult(
            source_path=path,
            source_type=canonical_type,
            headers=[],
            rows=[],
            header_row_index=None,
            sheet_name=None,
            extraction_notes=["ERROR: openpyxl not installed."],
        )]

    results: List[ExtractionResult] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        sheets_to_read = [target_sheet] if target_sheet else sheet_names
        for sheet_name in sheets_to_read:
            if sheet_name not in sheet_names:
                continue
            ws = wb[sheet_name]
            raw_rows: List[List[Any]] = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append(list(row))

            notes: List[str] = []
            header_idx, headers = locate_header_row(raw_rows, known_aliases=known_aliases)
            if header_idx > 0:
                notes.append(
                    f"Header detected at row {header_idx + 1} (1-based) in sheet '{sheet_name}'."
                )

            data_rows = extract_data_rows(raw_rows, header_idx, headers)
            results.append(ExtractionResult(
                source_path=path,
                source_type=canonical_type,
                headers=headers,
                rows=data_rows,
                header_row_index=header_idx,
                sheet_name=sheet_name,
                extraction_notes=notes,
            ))

        wb.close()
    except Exception as e:
        logger.error(f"XLSX read failed for {path}: {e}")
        results.append(ExtractionResult(
            source_path=path,
            source_type=canonical_type,
            headers=[],
            rows=[],
            header_row_index=None,
            sheet_name=None,
            extraction_notes=[f"ERROR: {e}"],
        ))
    return results


def get_excel_sheet_names(path: str) -> List[str]:
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []
